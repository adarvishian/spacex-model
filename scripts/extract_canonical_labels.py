#!/usr/bin/env python3
"""One-time extraction of column-A canonical labels from V2.16 into canonical_labels.py."""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = (
    REPO_ROOT
    / "Pre Existing Model Package"
    / "01_Current_State"
    / "SpaceX Model V2.16.xlsx"
)
OUTPUT_PATH = REPO_ROOT / "src" / "spacex_model" / "config" / "canonical_labels.py"
SKIP_SHEETS = frozenset({"Claude Log"})


def label_to_constant_name(label: str) -> str:
    """Convert Excel column-A label to a valid Python identifier."""
    name = label.upper()
    name = re.sub(r"[^A-Z0-9]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if name and name[0].isdigit():
        name = f"L_{name}"
    if not name:
        name = "UNNAMED"
    return name


def extract_labels(workbook_path: Path) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    by_sheet: dict[str, list[str]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        labels: list[str] = []
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=1, max_row=400, min_col=1, max_col=1):
            value = row[0].value
            if not value or not isinstance(value, str):
                continue
            text = value.strip()
            if len(text) < 2 or text.startswith("§"):
                continue
            if text not in seen:
                seen.add(text)
                labels.append(text)
        by_sheet[sheet_name] = labels
    wb.close()
    return by_sheet


def render_module(by_sheet: dict[str, list[str]], workbook_path: Path) -> str:
    all_labels: list[str] = []
    for labels in by_sheet.values():
        all_labels.extend(labels)
    unique_sorted = sorted(set(all_labels))

    lines = [
        '"""Canonical label registry — append-only per PRD §2.4 / Rule 10.',
        "",
        f"Extracted from: {workbook_path.name}",
        "Regenerate via: python scripts/extract_canonical_labels.py",
        '"""',
        "",
        "from __future__ import annotations",
        "",
        "from typing import Final",
        "",
        "# fmt: off",
        "",
    ]

    used_names: dict[str, str] = {}
    for label in unique_sorted:
        const = label_to_constant_name(label)
        base = const
        n = 2
        while const in used_names and used_names[const] != label:
            const = f"{base}_{n}"
            n += 1
        used_names[const] = label
        escaped = label.replace("\\", "\\\\").replace('"', '\\"')
        lines.append(f'{const}: Final[str] = "{escaped}"')

    lines.extend(
        [
            "",
            "# fmt: on",
            "",
            "CANONICAL_LABELS: Final[frozenset[str]] = frozenset(",
            "    {",
        ]
    )
    for label in unique_sorted:
        escaped = label.replace("\\", "\\\\").replace('"', '\\"')
        lines.append(f'        "{escaped}",')
    lines.extend(["    }", ")", "", "LABELS_BY_SHEET: Final[dict[str, tuple[str, ...]]] = {"])
    for sheet, labels in sorted(by_sheet.items()):
        sheet_key = sheet.replace("'", "\\'")
        lines.append(f'    "{sheet_key}": (')
        for label in labels:
            escaped = label.replace("\\", "\\\\").replace('"', '\\"')
            lines.append(f'        "{escaped}",')
        lines.append("    ),")
    lines.append("}")
    lines.append("")
    lines.append("")
    lines.append("def resolve_label(constant_name: str) -> str:")
    lines.append('    """Return the Excel label string for a registry constant name."""')
    lines.append("    return globals()[constant_name]")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 1
    by_sheet = extract_labels(workbook)
    total = sum(len(v) for v in by_sheet.values())
    print(f"Extracted {total} labels across {len(by_sheet)} sheets")
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(render_module(by_sheet, workbook), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
