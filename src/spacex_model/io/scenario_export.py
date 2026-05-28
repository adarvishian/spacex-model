"""Client Mode xlsx exports — active scenario and scenario pack (FRONTEND_PRD §7.4)."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from spacex_model.config.constants import FIRST_YEAR
from spacex_model.engine.pipeline import ModelResult, run_pipeline
from spacex_model.inputs.scenarios import load_scenario
from spacex_model.service.client_config import (
    CLIENT_INPUT_SPECS,
    client_overrides_to_canonical,
)
from spacex_model.service.serializers import _MODULE_DISPLAY, _MODULE_KEYS, _YEARS

_MODULE_SHEET_SLUG = {
    "customer_launch": "customer_launch",
    "starlink": "starlink",
    "odc": "odc",
    "ai_stack": "ai_stack",
    "lunar_mars": "lunar_mars",
}


def _audit_cell_url(base_url: str, sheet_slug: str, row_id: str, year: int) -> str:
    return f"{base_url.rstrip('/')}/audit/{sheet_slug}?row={row_id}&col={year}"


def _write_year_header(ws: Worksheet, col_start: int = 2) -> None:
    for i, year in enumerate(_YEARS):
        ws.cell(row=1, column=col_start + i, value=year)


def _write_series_row(
    ws: Worksheet,
    row: int,
    label: str,
    values: list[float],
    *,
    base_url: str | None = None,
    sheet_slug: str | None = None,
    row_id: str | None = None,
) -> None:
    ws.cell(row=row, column=1, value=label)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=2 + i, value=round(v, 2) if v is not None else None)
        if base_url and sheet_slug and row_id:
            year = _YEARS[i]
            cell.hyperlink = _audit_cell_url(base_url, sheet_slug, row_id, year)
            cell.font = Font(color="0563C1", underline="single")


def _module_rows(mod: Any) -> list[tuple[str, Any, str]]:
    return [
        ("Total Revenue ($mm)", mod.total_revenue, "R8"),
        ("Module EBITDA ($mm)", mod.module_ebitda, "R10"),
        ("Module FCF ($mm)", mod.module_fcf, "R14"),
    ]


def _write_module_sheet(
    wb: Workbook,
    key: str,
    result: ModelResult,
    *,
    base_url: str | None,
) -> None:
    name = _MODULE_DISPLAY[key]
    ws = wb.create_sheet(name[:31])
    slug = _MODULE_SHEET_SLUG[key]
    mod = result.module_outputs[key]
    _write_year_header(ws)
    for r, (lbl, series, row_id) in enumerate(_module_rows(mod), start=2):
        vals = [float(series.at(y)) for y in _YEARS]
        _write_series_row(ws, r, lbl, vals, base_url=base_url, sheet_slug=slug, row_id=row_id)


def _write_group_sheet(wb: Workbook, result: ModelResult, *, base_url: str | None) -> None:
    ws = wb.create_sheet("Group")
    g = result.group_pnl
    _write_year_header(ws)
    rows = [
        ("Group Revenue (net) ($mm)", g.group_revenue_net, "R8"),
        ("Group EBITDA ($mm)", g.group_ebitda, "R10"),
        ("Group FCF ($mm)", g.group_fcf, "R102"),
    ]
    for r, (lbl, series, row_id) in enumerate(rows, start=2):
        vals = [float(series.at(y)) for y in _YEARS]
        _write_series_row(
            ws, r, lbl, vals, base_url=base_url, sheet_slug="group_pnl", row_id=row_id
        )


def _write_valuation_sheet(wb: Workbook, result: ModelResult, *, base_url: str | None) -> None:
    ws = wb.create_sheet("Valuation")
    val = result.valuation
    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    metrics = [
        ("Group EV 2025 ($B)", val.implied_ev_2025_billions),
        ("Group WACC", val.group_wacc),
        ("Terminal growth", val.terminal_growth),
    ]
    for r, (lbl, v) in enumerate(metrics, start=2):
        ws.cell(row=r, column=1, value=lbl)
        c = ws.cell(row=r, column=2, value=round(float(v), 4))
        if base_url:
            c.hyperlink = _audit_cell_url(base_url, "valuation", "D6", FIRST_YEAR)
            c.font = Font(color="0563C1", underline="single")


def _write_cover(
    wb: Workbook,
    *,
    scenario_name: str,
    description: str,
    overrides: dict[str, Any],
    run_id: str,
    group_ev_b: float,
    base_url: str | None,
) -> None:
    ws = wb.create_sheet("Cover", 0)
    lines = [
        ("Mach33 SpaceX Valuation Model", None),
        ("Scenario", scenario_name),
        ("Description", description),
        ("Run ID", run_id),
        ("Group EV (2025)", f"${group_ev_b:,.1f}B"),
        ("", None),
        ("Scenario inputs", None),
    ]
    row = 1
    for a, b in lines:
        ws.cell(row=row, column=1, value=a)
        if b is not None:
            ws.cell(row=row, column=2, value=b)
        row += 1
    for spec in CLIENT_INPUT_SPECS:
        label = spec["canonical_label"]
        if label in overrides:
            ws.cell(row=row, column=1, value=spec["plain_label"])
            ws.cell(row=row, column=2, value=overrides[label])
            row += 1
    if base_url:
        ws.cell(row=row + 1, column=1, value="Audit derivations")
        ws.cell(row=row + 1, column=2, value=f"{base_url.rstrip('/')}/audit")
        link = ws.cell(row=row + 1, column=2)
        link.hyperlink = f"{base_url.rstrip('/')}/audit"
        link.font = Font(color="0563C1", underline="single")


def export_active_scenario_xlsx(
    result: ModelResult,
    *,
    scenario_name: str,
    description: str,
    base_url: str = "",
) -> bytes:
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.active)
    _write_cover(
        wb,
        scenario_name=scenario_name,
        description=description,
        overrides=result.audit.get("overrides") or {},
        run_id=result.run_id,
        group_ev_b=result.valuation.implied_ev_2025_billions,
        base_url=base_url or None,
    )
    for key in _MODULE_KEYS:
        _write_module_sheet(wb, key, result, base_url=base_url or None)
    _write_group_sheet(wb, result, base_url=base_url or None)
    _write_valuation_sheet(wb, result, base_url=base_url or None)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_scenario_pack_xlsx(
    results: dict[str, ModelResult],
    descriptions: dict[str, str],
    *,
    base_url: str = "",
) -> bytes:
    """Base / Bear / Bull columns per metric row on shared module sheets."""
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.active)
    ws_cover = wb.create_sheet("Cover", 0)
    ws_cover.cell(row=1, column=1, value="Mach33 SpaceX Valuation — Scenario Pack")
    ws_cover.cell(row=2, column=1, value="Scenarios")
    ws_cover.cell(row=2, column=2, value=", ".join(results.keys()))
    if base_url:
        ws_cover.cell(row=3, column=1, value="Audit")
        c = ws_cover.cell(row=3, column=2, value=base_url.rstrip("/") + "/audit")
        c.hyperlink = base_url.rstrip("/") + "/audit"
        c.font = Font(color="0563C1", underline="single")

    scenario_ids = list(results.keys())
    for key in _MODULE_KEYS:
        name = _MODULE_DISPLAY[key]
        ws = wb.create_sheet(name[:31])
        col = 1
        ws.cell(row=1, column=col, value="Metric / Year")
        col += 1
        for sid in scenario_ids:
            ws.cell(row=1, column=col, value=sid.replace("_", " ").title())
            col += len(_YEARS) + 1

        row = 2
        for lbl, _series, _row_id in _module_rows(results[scenario_ids[0]].module_outputs[key]):
            ws.cell(row=row, column=1, value=lbl)
            col = 2
            for sid in scenario_ids:
                mod = results[sid].module_outputs[key]
                ser = next(s for l2, s, _ in _module_rows(mod) if l2 == lbl)
                for i, year in enumerate(_YEARS):
                    ws.cell(row=row, column=col + i, value=round(float(ser.at(year)), 2))
                col += len(_YEARS) + 1
            row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_for_export(
    scenario_path: Any,
    *,
    extra_overrides: dict[str, Any] | None = None,
) -> ModelResult:
    return run_pipeline(
        scenario_path=scenario_path,
        extra_overrides=extra_overrides,
        write_outputs=False,
    )
