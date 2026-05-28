"""Excel workbook ingest — formula pass + value pass per context.md §4.2."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from spacex_model.config.constants import (
    COL_BASE_CASE,
    COL_MC_DISTRIBUTION,
    COL_MC_MAX,
    COL_MC_MIN,
    COL_MC_NOTES,
    COL_NOTES,
    FIRST_YEAR,
    LAST_YEAR,
)


@dataclass(frozen=True, slots=True)
class CellRecord:
    sheet: str
    row: int
    col: int
    label: str | None
    value: Any
    formula: str | None


@dataclass(frozen=True, slots=True)
class AssumptionRowRecord:
    row: int
    label: str
    section: str
    base_case: float | str | None
    year_values: dict[int, float | str | None]
    notes: str | None
    mc_min: float | str | None
    mc_max: float | str | None
    distribution: str | None
    mc_notes: str | None


@dataclass
class FormulaPassResult:
    workbook_path: Path
    sheet_names: list[str]
    formulas: list[CellRecord] = field(default_factory=list)
    labels_by_sheet: dict[str, dict[int, str]] = field(default_factory=dict)


@dataclass
class ValuePassResult:
    workbook_path: Path
    cached_values: dict[tuple[str, int, int], Any] = field(default_factory=dict)
    assumptions_rows: list[AssumptionRowRecord] = field(default_factory=list)
    labels_by_sheet: dict[str, dict[int, str]] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class DemandCurveBreakpoint:
    """Single Q→Revenue row from Demand Curves tab."""

    label: str
    quantity_gbps: float
    revenue: float


@dataclass
class DemandCurvesPassResult:
    """Breakpoint tables parsed from Demand Curves tab."""

    bb_breakpoints: list[DemandCurveBreakpoint] = field(default_factory=list)
    dtc_breakpoints: list[DemandCurveBreakpoint] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class IngestResult:
    workbook_path: Path
    formula_pass: FormulaPassResult
    value_pass: ValuePassResult
    demand_curves: DemandCurvesPassResult | None = None


def _detect_year_columns(ws: Worksheet, scan_rows: int = 6) -> dict[int, int]:
    """Map calendar year -> 1-based column index from header row."""
    for row_idx in range(1, scan_rows + 1):
        mapping: dict[int, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx).value
            if isinstance(cell, (int, float)) and FIRST_YEAR <= int(cell) <= LAST_YEAR:
                mapping[int(cell)] = col_idx
        if len(mapping) >= 10:
            return mapping
    return {}


def _current_section(label: str, section: str) -> str:
    if label.startswith("§"):
        return label.strip()
    return section


def _coerce_number(value: Any) -> float | str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return text
    return str(value)


def _extract_labels(ws: Worksheet, max_row: int = 400) -> dict[int, str]:
    labels: dict[int, str] = {}
    for row_idx in range(1, max_row + 1):
        value = ws.cell(row_idx, 1).value
        if value and isinstance(value, str) and len(value.strip()) > 1:
            labels[row_idx] = value.strip()
    return labels


def run_formula_pass(workbook_path: Path) -> FormulaPassResult:
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    result = FormulaPassResult(workbook_path=workbook_path, sheet_names=list(wb.sheetnames))
    for sheet_name in wb.sheetnames:
        if sheet_name == "Claude Log":
            continue
        ws = wb[sheet_name]
        labels = _extract_labels(ws)
        result.labels_by_sheet[sheet_name] = labels
        for row_idx, label in labels.items():
            for col_idx in range(1, min(ws.max_column, 40) + 1):
                cell = ws.cell(row_idx, col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result.formulas.append(
                        CellRecord(
                            sheet=sheet_name,
                            row=row_idx,
                            col=col_idx,
                            label=label,
                            value=None,
                            formula=cell.value,
                        )
                    )
    wb.close()
    return result


def run_value_pass(workbook_path: Path) -> ValuePassResult:
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    result = ValuePassResult(workbook_path=workbook_path)
    if "Assumptions" not in wb.sheetnames:
        wb.close()
        msg = "Workbook missing Assumptions tab"
        raise ValueError(msg)

    assumptions_ws = wb["Assumptions"]
    year_cols = _detect_year_columns(assumptions_ws)
    if not year_cols:
        result.warnings.append("Could not detect year header row on Assumptions tab")

    section = "§unknown"
    for row_idx in range(1, assumptions_ws.max_row + 1):
        label_val = assumptions_ws.cell(row_idx, 1).value
        if not label_val or not isinstance(label_val, str):
            continue
        label = label_val.strip()
        if not label:
            continue
        section = _current_section(label, section)
        if label.startswith("§"):
            continue

        base_case = _coerce_number(assumptions_ws.cell(row_idx, COL_BASE_CASE).value)
        year_values: dict[int, float | str | None] = {}
        for year, col_idx in year_cols.items():
            year_values[year] = _coerce_number(assumptions_ws.cell(row_idx, col_idx).value)

        if base_case is None and not any(v is not None for v in year_values.values()):
            continue

        mc_dist_val = assumptions_ws.cell(row_idx, COL_MC_DISTRIBUTION).value
        distribution = str(mc_dist_val).strip() if mc_dist_val else None

        result.assumptions_rows.append(
            AssumptionRowRecord(
                row=row_idx,
                label=label,
                section=section,
                base_case=base_case,
                year_values=year_values,
                notes=(
                    str(assumptions_ws.cell(row_idx, COL_NOTES).value)
                    if assumptions_ws.cell(row_idx, COL_NOTES).value
                    else None
                ),
                mc_min=_coerce_number(assumptions_ws.cell(row_idx, COL_MC_MIN).value),
                mc_max=_coerce_number(assumptions_ws.cell(row_idx, COL_MC_MAX).value),
                distribution=distribution,
                mc_notes=(
                    str(assumptions_ws.cell(row_idx, COL_MC_NOTES).value)
                    if assumptions_ws.cell(row_idx, COL_MC_NOTES).value
                    else None
                ),
            )
        )

    for sheet_name in wb.sheetnames:
        if sheet_name == "Claude Log":
            continue
        ws = wb[sheet_name]
        labels = _extract_labels(ws)
        result.labels_by_sheet[sheet_name] = labels
        year_cols_sheet = _detect_year_columns(ws)
        for row_idx, label in labels.items():
            for year, col_idx in year_cols_sheet.items():
                result.cached_values[(sheet_name, row_idx, year)] = _coerce_number(
                    ws.cell(row_idx, col_idx).value
                )
            for col_idx in range(1, min(ws.max_column, 35) + 1):
                if col_idx in year_cols_sheet.values():
                    continue
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    result.cached_values[(sheet_name, row_idx, col_idx)] = val

    wb.close()
    return result


def _parse_demand_curves_pass(workbook_path: Path) -> DemandCurvesPassResult:
    """Extract BB/DTC piecewise-linear breakpoints from Demand Curves tab."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    result = DemandCurvesPassResult()
    if "Demand Curves" not in wb.sheetnames:
        result.warnings.append("Workbook missing Demand Curves tab")
        wb.close()
        return result

    ws = wb["Demand Curves"]
    for row_idx in range(1, ws.max_row + 1):
        label_val = ws.cell(row_idx, 1).value
        if not label_val or not isinstance(label_val, str):
            continue
        label = label_val.strip()
        q_val = _coerce_number(ws.cell(row_idx, 2).value)
        rev_val = _coerce_number(ws.cell(row_idx, 3).value)
        if q_val is None or rev_val is None:
            continue
        if label.startswith("BB Q breakpoint"):
            result.bb_breakpoints.append(
                DemandCurveBreakpoint(label=label, quantity_gbps=float(q_val), revenue=float(rev_val))
            )
        elif label.startswith("DTC Q breakpoint"):
            result.dtc_breakpoints.append(
                DemandCurveBreakpoint(label=label, quantity_gbps=float(q_val), revenue=float(rev_val))
            )

    if not result.bb_breakpoints:
        result.warnings.append("No BB Q breakpoints found on Demand Curves tab")
    if not result.dtc_breakpoints:
        result.warnings.append("No DTC Q breakpoints found on Demand Curves tab")
    wb.close()
    return result


def ingest_workbook(workbook_path: Path) -> IngestResult:
    """Run formula pass then value pass."""
    formula = run_formula_pass(workbook_path)
    values = run_value_pass(workbook_path)
    demand = _parse_demand_curves_pass(workbook_path)
    values.warnings.extend(demand.warnings)
    return IngestResult(
        workbook_path=workbook_path,
        formula_pass=formula,
        value_pass=values,
        demand_curves=demand,
    )
